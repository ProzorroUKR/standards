import cloudscraper
import json
import re
import requests

from bs4 import BeautifulSoup
from io import BytesIO
from openpyxl import load_workbook

MINDEV_RESOURCE = "https://mindev.gov.ua"
KATOTTG_RESOURCE = f"{MINDEV_RESOURCE}/diialnist/rozvytok-mistsevoho-samovriaduvannia/kodyfikator-administratyvno-terytorialnykh-odynyts-ta-terytorii-terytorialnykh-hromad"


def get_katottg():
    scraper = cloudscraper.create_scraper()  # emulate browser
    response = scraper.get(KATOTTG_RESOURCE)
    if response.status_code == 200:
        soup = BeautifulSoup(response.text, 'html.parser')
        katottg_link = soup.find("a", attrs={"class": "fr-file", "data-extension": "xlsx"}).get("href")
        if katottg_link:
            export_katottg(katottg_link)
        else:
            print("KATOTTG link not found")
    else:
        print(f"Response from resource {MINDEV_RESOURCE}: {response.status_code} - {response.text}")


def export_katottg(katottg_link):
    response = requests.get(katottg_link)
    if response.status_code == 200:
        katottg_dict = {}
        workbook = load_workbook(BytesIO(response.content))
        # Active sheet in excel
        worksheet = workbook.active
        for row in worksheet.iter_rows(values_only=True):
            try:
                if row[0] and re.match(r"^UA\d{17}$", row[0]):
                    idx = find_katottg_code_index(row)
                    if idx >= 0:
                        katottg_dict[row[idx]] = {
                            "name":  row[-1],
                            "category": row[-2],
                            "level": idx + 1,
                            "parent_code": row[idx-1] if idx - 1 >= 0 else "",
                        }
            except Exception as e:
                print(f"ERROR: caught {type(e).__name__}.")

        if katottg_dict:
            sorted_data = dict(sorted(katottg_dict.items()))
            json_object = json.dumps(sorted_data, indent=2, ensure_ascii=False)
            with open(f'./classifiers/katottg.json', 'w') as outfile:
                outfile.write(json_object)
                outfile.write("\n")  # empty line in the end of file for json validation
        else:
            print(f"Something went wrong and there is no data fetched from resource {MINDEV_RESOURCE}")
    else:
        print(f"Response from resource {MINDEV_RESOURCE} during loading file: {response.status_code} - {response.text}")


def find_katottg_code_index(row):
    i = len(row) - 3  # get last cell with code in row
    while i >= 0 and (not row[i] or not row[i].startswith("UA")):
        i -= 1

    if i >= 0:
        return i


if __name__ == "__main__":
    get_katottg()
